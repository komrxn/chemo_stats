"""
Excel Export Service
Generates Excel files with ANOVA results (5 sheets, like MATLAB output)
"""
import logging
from io import BytesIO

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)


def generate_anova_excel(
    results: list[dict],
    multicomp: list[dict],
    global_stats: dict,
    group_stats: dict,
    data: np.ndarray,
    classes: np.ndarray,
    var_names: list[str]
) -> BytesIO:
    """
    Generate Excel file with ANOVA results (5 sheets)
    
    Sheets:
    1. ANOVA_TABLE_KKH - Main results table
    2. MULTICOMPARISON - Pairwise comparisons
    3. DATASET - Original data
    4. GLOBALSTATDATA - Global statistics
    5. GROUPSTATDATA - Group statistics
    
    Args:
        results: ANOVA results table
        multicomp: Multicomparison results
        global_stats: Global statistics
        group_stats: Group statistics
        data: Original data matrix
        classes: Class labels
        var_names: Variable names
    
    Returns:
        BytesIO object with Excel file
    """
    logger.info("Generating Excel file...")
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Sheet 1: ANOVA_TABLE_KKH
    _create_anova_table_sheet(wb, results)
    
    # Sheet 2: MULTICOMPARISON
    _create_multicomp_sheet(wb, multicomp)
    
    # Sheet 3: DATASET
    _create_dataset_sheet(wb, data, classes, var_names)
    
    # Sheet 4: GLOBALSTATDATA
    _create_global_stats_sheet(wb, global_stats)
    
    # Sheet 5: GROUPSTATDATA
    _create_group_stats_sheet(wb, group_stats, var_names)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info("✅ Excel file generated successfully")
    
    return output


def _create_anova_table_sheet(wb: Workbook, results: list[dict]):
    """Sheet 1: ANOVA_TABLE_KKH"""
    ws = wb.create_sheet("ANOVA_TABLE_KKH", 0)
    
    # Header row
    headers = ['VariableIndex', 'Variable', 'P-Nominal', 'P_FDR', 'Effect size (%)', 'F-stat']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for i, result in enumerate(results):
        ws.append([
            i + 1,
            result['variable'],
            result['pValue'],
            result['fdr'],
            result['effectSize'],
            result['fStat']
        ])
        
        # Highlight significant variables (Benjamini)
        row_idx = i + 2
        if result['benjamini']:
            for cell in ws[row_idx]:
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_multicomp_sheet(wb: Workbook, multicomp: list[dict]):
    """Sheet 2: MULTICOMPARISON"""
    ws = wb.create_sheet("MULTICOMPARISON")
    
    # Header row
    headers = ['VariableIndex', 'Variable', 'GroupX', 'GroupY', 'P_Nominal', 'P_FDR', 'MeanDiff', 'T-stat']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for comp in multicomp:
        ws.append([
            comp['variableIndex'],
            comp['variable'],
            comp['groupX'],
            comp['groupY'],
            comp['pValue'],
            comp.get('pValue_FDR', comp['pValue']),
            comp['mean_diff'],
            comp.get('tStat', 0.0)
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_dataset_sheet(wb: Workbook, data: np.ndarray, classes: np.ndarray, var_names: list[str]):
    """Sheet 3: DATASET"""
    ws = wb.create_sheet("DATASET")
    
    # First row: empty, "Variable#", then variable numbers
    ws.append(['', 'Variable#'] + list(range(1, len(var_names) + 1)))
    
    # Second row: "Sample#", "Design", then variable names
    ws.append(['Sample#', 'Design'] + var_names)
    
    # Data rows
    for i in range(len(data)):
        ws.append([i + 1, int(classes[i])] + data[i].tolist())
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font


def _create_global_stats_sheet(wb: Workbook, global_stats: dict):
    """Sheet 4: GLOBALSTATDATA"""
    ws = wb.create_sheet("GLOBALSTATDATA")
    
    # Header row
    headers = ['Variable', 'RSD', 'STD', 'MEAN', 'RANGE', 'MIN', 'MAX']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for i, var in enumerate(global_stats['variables']):
        ws.append([
            var,
            global_stats['RSD'][i],
            global_stats['STD'][i],
            global_stats['MEAN'][i],
            global_stats['RANGE'][i],
            global_stats['MIN'][i],
            global_stats['MAX'][i]
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _create_group_stats_sheet(wb: Workbook, group_stats: dict, var_names: list[str]):
    """Sheet 5: GROUPSTATDATA"""
    ws = wb.create_sheet("GROUPSTATDATA")
    
    # Create headers: Variable, then for each group: RSD, STD, MEAN, RANGE, MIN, MAX
    headers = ['Variable']
    for group_name in sorted(group_stats.keys()):
        headers.extend([
            f'{group_name}-RSD',
            f'{group_name}-STD',
            f'{group_name}-MEAN',
            f'{group_name}-RANGE',
            f'{group_name}-MIN',
            f'{group_name}-MAX'
        ])
    
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for i, var in enumerate(var_names):
        row = [var]
        for group_name in sorted(group_stats.keys()):
            stats = group_stats[group_name]
            row.extend([
                stats['RSD'][i],
                stats['STD'][i],
                stats['MEAN'][i],
                stats['RANGE'][i],
                stats['MIN'][i],
                stats['MAX'][i]
            ])
        ws.append(row)
    
    # Auto-adjust column widths (limit to avoid very wide columns)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width






